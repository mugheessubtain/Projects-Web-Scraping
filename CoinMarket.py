from selenium import webdriver
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.common.by import By  
import time

path = r"C:\Users\hp\Downloads\chromedriver-win64\chromedriver-win64\chromedriver.exe"
data = []

options = Options()
# options.add_argument("--headless")
for i in range(1,3): ## from 1 to number of pages in the website
    driver = webdriver.Chrome(path, options=options)
    driver.get(f"https://coinmarketcap.com/?page={i}")

    scroll_pause_time = 0.5  # increased pause for slower scroll
    scroll_increment = 300  # smaller scroll step for smoother movement
    current_position = 0
    max_scroll = driver.execute_script("return document.body.scrollHeight")

    while current_position < max_scroll:
        driver.execute_script(f"window.scrollTo(0, {current_position});")
        time.sleep(scroll_pause_time)
        current_position += scroll_increment
        max_scroll = driver.execute_script("return document.body.scrollHeight")
    import pandas as pd

    table = driver.find_element(By.CLASS_NAME , "cmc-table")
    rows = table.find_elements(By.TAG_NAME, "tr")
    print(rows[0].text.replace('\n', ' | '))
    columns = rows[0].text
    for row in rows[1:]:
        table_data = row.find_elements(By.TAG_NAME , "td")
        rank = table_data[1].text
        name = table_data[2].text
        hour_percent = table_data[3].text
        twentyFour_hour_percent = table_data[4].text
        seven_day_percent = table_data[5].text
        market_cap = table_data[6].text
        volume_24h = table_data[7].text
        c_supply = table_data[8].text
    
        print(rank,name,hour_percent,twentyFour_hour_percent,seven_day_percent,market_cap,volume_24h,c_supply)
    
        data.append({
            "rank":rank ,
            "name":name,
            "hour_percent":hour_percent,                
            "twentyFour_hour_percent":twentyFour_hour_percent,
            "seven_day_percent":seven_day_percent,
            "market_cap":market_cap,
            "volume_24h":volume_24h,
            "c_supply":c_supply
        
            })


df = pd.DataFrame(data)
import os
from openpyxl import load_workbook

file_path = "coinmarketcap.xlsx"

# Check if the file exists first
if not os.path.exists(file_path):
    # If file doesn't exist, write normally with headers
    df.to_excel(file_path, index=False)
else:
    # If file exists, append without headers
    book = load_workbook(file_path)
    sheet = book.active
    startrow = sheet.max_row

    with pd.ExcelWriter(file_path, engine='openpyxl', mode='a', if_sheet_exists='overlay') as writer:
        df.to_excel(writer, index=False, header=False, startrow=startrow)

